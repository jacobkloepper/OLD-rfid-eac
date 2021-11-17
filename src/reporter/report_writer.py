from openpyxl import Workbook
from openpyxl import load_workbook

import time
import util
import os.path
from calendar import monthrange
import db_interface as dbi
import log_interface as li
import dates


# Calling the function Reporter.update_ytd() method will fill in each day from the last recorded day until today 
#   with the data found in each day's log file.

class ReportWriter:

    def __init__(self, **kwargs):
        self.date = time.localtime(time.time())
        self.datafile = "reports/attendance/{}.xlsx".format(self.date.tm_year)
        self.output_file = "default.xlsx"
        self.pop = False
        self._DEBUG = True

        # kwargs resolution
        if (kwargs.get("path")):
            self.datafile = kwargs.get("path")
        if (kwargs.get("output")):
            self.output_file = kwargs.get("output")
        if (kwargs.get("pop")):
            self.pop = kwargs.get("pop")


        # Write Write pop or attendance log
        self.wb_file = self.datafile

        # Create / setup workbook
        if (not os.path.isfile(self.wb_file)):
            # Create workbook and set working sheet
            if (self.pop == False):
                self.create_report()
                self.ws = self.wb["{} {}".format(util.mons.get(self.date.tm_mon), self.date.tm_year)]
            else:
                self.create_pop()
                self.ws = self.wb["IOU"]
            print("File created at {}".format(self.wb_file))
        else:
            self.wb = load_workbook(self.wb_file)
            if (self.pop == False):
                self.ws = self.wb["{} {}".format(util.mons.get(self.date.tm_mon), self.date.tm_year)]
            else:
                self.ws = self.wb["IOU"]

        self.sheet_list = []
        for sheet in self.wb.sheetnames:
            self.sheet_list.append(self.wb[sheet])


    def create_report(self):
        self.wb = Workbook()

        month_names = []
        ws_list = []
        self.sheet_list = []

        # Create all sheets
        for i in range(12):
            month_names.append("{} {}".format(util.mons.get(i+1), self.date.tm_year))
            ws_list.append(self.wb.create_sheet(month_names[i]))

        # Remove the init sheet, set the active sheet to this month
        self.wb.remove_sheet(self.wb.get_sheet_by_name("Sheet"))
        for sheet in self.wb.sheetnames:
            self.sheet_list.append(self.wb[sheet])

        # Write header info (pop reports have different format)
        self.format_users()
        self.format_dates()

        self.wb.save(self.wb_file)
    def format_users(self):
    # Print the first and last names of all users found in the database
        db = dbi.db_interface()
    
        firsts = db.get_usernames()[0]
        lasts = db.get_usernames()[1]
        
        
        # Write first names
        for sheet in self.sheet_list:
            num = 0
            for col in sheet.iter_cols(min_row=4, max_row=4+util.NUM_PERSONS-1, min_col=1, max_col=1):
                for cell in col:
                    cell.value = firsts[num]
                    num = num + 1

        # Write last names
        for sheet in self.sheet_list:
            num = 0
            for col in sheet.iter_cols(min_row=4, max_row=4+util.NUM_PERSONS-1, min_col=2, max_col=2):
                for cell in col:
                    cell.value = lasts[num]
                    num = num + 1
    def format_dates(self):
        # Iteration vars (for each sheet/month)
        year = self.date.tm_year
        month = 1

        # Write company name
        for sheet in self.sheet_list:
            sheet["A1"].value = "ACME"
            sheet["B1"].value = "CORP"

        # Write month year
        month = 1
        for sheet in self.sheet_list:
            cell = sheet["C1"]
            cell.value = "{} {}".format(util.mons.get(month).upper(), year)
            month = month + 1

        # Write date nums
        month = 1
        for sheet in self.sheet_list:
            month_info = monthrange(self.date.tm_year, month)  # month_info = (d1, dl): d1 is the number of first weekday of month, dl is the number of days in month
            for row in sheet.iter_rows(min_row=2, max_row=2, min_col=3, max_col=3+month_info[1]-1):
                for cell in row:
                    cell.value = "{:02}".format(cell.col_idx-2)
            month = month + 1

        # Write wdays
        month = 1
        for sheet in self.sheet_list:
            month_info = monthrange(self.date.tm_year, month)
            weekday = month_info[0]
            for row in sheet.iter_rows(min_row=3, max_row=3, min_col=3, max_col=3+month_info[1]-1):
                for cell in row:
                    cell.value = util.wdays_abbrev.get(weekday)
                    weekday = (weekday + 1) % 7

            month = month + 1

    def find_next_day(self):
    # Scan the report file to find the most recently updated column.
        for sheet in self.sheet_list:
            for i in range(int(sheet.max_column)-3):
                for col in sheet.iter_cols(min_col=i+3, max_col=i+3, min_row=4, max_row=4+util.NUM_PERSONS-1):
                    if (util.all_of(col, None)):
                        title = sheet.title.split()
                        return (i+1, util.mon_num.get(title[0]), int(title[1]))
    def update_ytd(self):
        next_day = self.find_next_day()
        today = (self.date.tm_mday, self.date.tm_mon, self.date.tm_year)
        yesterday = dates.decrement(today)

        for date in dates.dates_between(next_day, yesterday):
            try:
                self.write_day(date)

                if self._DEBUG:
                    print("SUCCESS: write {}...".format(date))
            except:
                if self._DEBUG:
                    print("FAILED : write {}...".format(date))
    def write_day(self, date):
    # I feel like this function doesn't need to have a nested for
    # Also, needs to grab only the outermost enter-exit pair.
        log = li.log_interface()
        day_info = log.get_day_log(date)

        update_sheet = self.wb["{} {}".format(util.mons.get(date[1]), date[2])]

        # Scan through all rows
        for row in update_sheet[4:4+util.NUM_PERSONS-1]:
            # For each row, look for the corresponding user log
            for listing in day_info:
                # If we get a match, fill in
                if (row[0].value == listing[0] and row[1].value == listing[1]):
                    if (listing[3] == "enter"):
                        row[date[0]+1].value = listing[2]
                    else:
                        row[date[0]+1].value = "{} - {}".format(row[date[0]+1].value, listing[2])

    
    def create_pop(self):
        self.wb = Workbook()
        self.sheet_list = []

        self.wb.create_sheet("IOU")
        self.wb.remove_sheet(self.wb.get_sheet_by_name("Sheet"))

        for sheet in self.wb.sheetnames:
            self.sheet_list.append(self.wb[sheet])

        self.format_users()
        self.format_pop()
    def format_pop(self):
        for sheet in self.sheet_list:
            sheet['C3'] = "IOU quarters"
    def update_pop(self):
        today = (self.date.tm_mday, self.date.tm_mon, self.date.tm_year)
        self.write_pop(today)
    def write_pop(self, date):
        log = li.log_interface()
        day_info = log.get_pop_log(date)
        day = date[0]
        month = date[1]
        year = date[2]

        update_sheet = self.wb["IOU"]

        # Scan through all rows
        for row in update_sheet[4:4+util.NUM_PERSONS-1]:
            # For each row, look for the corresponding user log
            for listing in day_info:
                # If we get a match, fill in
                if (row[0].value == listing[0] and row[1].value == listing[1]):
                    if (row[2].value != None):
                        row[2].value = row[2].value+1
                    else:
                        row[2].value = 1

    def checkoff(self, firstname, lastname, quantity):
        update_sheet = self.wb["IOU"]

        for row in update_sheet[4:4+util.NUM_PERSONS-1]:
        # For each row, look for the corresponding user log
            if (row[0].value == firstname and row[1].value == lastname):
                if (row[2].value != None):
                    if (quantity == "all"):
                        row[2].value = 0
                    else:
                        row[2].value = row[2].value-int(quantity)
                    if (row[2].value < 0):
                        row[2].value = 0
                        print("Checked off more than user [{} {}]'s IOU count. Set to zero.".format(firstname,lastname))
                else:
                    print("User [{} {}] has no IOUs.".format(firstname, lastname))